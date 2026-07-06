"""
Build TMC_Vision_Accuracy_Summary.xlsx from the tidy CSVs / stats produced
by analyze_tmc.py: per-sheet data tabs, per-sheet stats tabs, a Baseline vs.
1st Calibration comparison tab, and native Excel charts.
"""

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SHEETS = ["Baseline", "1st Calibration"]
FILE_MAP = {"Baseline": "tmc_Baseline.csv", "1st Calibration": "tmc_1st_Calibration.csv"}
TAB_MAP = {"Baseline": "Baseline Data", "1st Calibration": "Calibration1 Data"}
STATS_TAB_MAP = {"Baseline": "Baseline Stats", "1st Calibration": "Calibration1 Stats"}

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, ncols):
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        max_len = max((len(str(ws.cell(row=r, column=c).value)) for r in range(1, ws.max_row + 1)
                       if ws.cell(row=r, column=c).value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


def write_data_tab(wb, sheet_name, df, stats):
    ws = wb.create_sheet(TAB_MAP[sheet_name])
    cols = ["ID", "Intersection", "Date", "Time of Day", "Manual_Total", "Vision_Total",
            "Pct_Diff", "GEH_15min", "GEH_Hourly_Equiv"]
    ws.append(cols)
    for _, row in df.iterrows():
        ws.append([row[c] for c in cols])
    style_header(ws, len(cols))
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=7).number_format = "0.0%"
        ws.cell(row=r, column=8).number_format = "0.00"
        ws.cell(row=r, column=9).number_format = "0.00"
    ws.freeze_panes = "A2"
    autosize(ws, len(cols))

    # Bar chart: GEH (15-min) per intersection, sorted (data already sorted desc by GEH)
    chart = BarChart()
    chart.title = f"{sheet_name}: GEH (15-min) by Intersection"
    chart.y_axis.title = "GEH"
    chart.x_axis.title = "Intersection"
    chart.height, chart.width = 10, 24
    data_ref = Reference(ws, min_col=8, min_row=1, max_row=ws.max_row)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None
    ws.add_chart(chart, "K1")
    note1 = ws.cell(row=21, column=11,
                     value="Sorted highest to lowest GEH. Conventional threshold is GEH < 5 (on hourly "
                           "volumes) — see the Overview tab for why that needs care on 15-minute counts.")
    note1.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=21, start_column=11, end_row=21, end_column=18)
    ws.row_dimensions[21].height = 28

    # Scatter: Manual vs Vision Total
    bias = stats["mean_pct_diff"]
    direction = "lower than (undercounted)" if bias < 0 else "higher than (overcounted)"
    note2 = ws.cell(row=22, column=11,
                     value=(f"How to read this: each point is one intersection — x is the manually counted "
                            f"Total, y is the Vision Camera's Total for the same window. Points on the 1:1 "
                            f"line are a perfect match; points below it mean Vision undercounted. On average, "
                            f"Vision's Total was {abs(bias) * 100:.1f}% {direction} the manual count in {sheet_name}."))
    note2.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=22, start_column=11, end_row=22, end_column=18)
    ws.row_dimensions[22].height = 42

    scatter = ScatterChart()
    scatter.title = f"{sheet_name}: Manual vs. Vision Total"
    scatter.x_axis.title = "Manual Total"
    scatter.y_axis.title = "Vision Total"
    scatter.height, scatter.width = 10, 14
    xvalues = Reference(ws, min_col=5, min_row=2, max_row=ws.max_row)
    yvalues = Reference(ws, min_col=6, min_row=2, max_row=ws.max_row)
    series = Series(yvalues, xvalues, title="Intersections")
    series.marker.symbol = "circle"
    series.graphicalProperties.line.noFill = True
    scatter.series.append(series)
    ws.add_chart(scatter, "K23")

    return ws


def write_stats_tab(wb, sheet_name, stats):
    ws = wb.create_sheet(STATS_TAB_MAP[sheet_name])
    labels = {
        "n_intersections": "Intersections with data",
        "mean_pct_diff": "Mean % Difference",
        "median_pct_diff": "Median % Difference",
        "mean_abs_pct_diff": "Mean |% Difference|",
        "mean_geh_15min": "Mean GEH (15-min, as collected)",
        "median_geh_15min": "Median GEH (15-min, as collected)",
        "mean_geh_hourly": "Mean GEH (hourly-equivalent, x4 volume)",
        "median_geh_hourly": "Median GEH (hourly-equivalent, x4 volume)",
        "pct_pass_geh5_15min": "% Intersections GEH<5 (15-min)",
        "pct_pass_geh5_hourly": "% Intersections GEH<5 (hourly-equivalent)",
    }
    ws.append(["Metric", "Value"])
    for key, label in labels.items():
        val = stats[key]
        if key.startswith("mean_pct") or key.startswith("median_pct") or key == "mean_abs_pct_diff":
            ws.append([label, val])
            ws.cell(row=ws.max_row, column=2).number_format = "0.0%"
        elif key.startswith("pct_pass"):
            ws.append([label, val / 100])
            ws.cell(row=ws.max_row, column=2).number_format = "0.0%"
        elif key == "n_intersections":
            ws.append([label, int(val)])
        else:
            ws.append([label, val])
            ws.cell(row=ws.max_row, column=2).number_format = "0.00"
    style_header(ws, 2)
    autosize(ws, 2)

    note = ws.cell(row=ws.max_row + 2, column=1,
                    value=("Note: counts are 15-minute samples. GEH scales with sqrt(volume), so the "
                           "conventional GEH<5 threshold (calibrated for hourly volumes) is stricter once "
                           "counts are scaled to an hourly-equivalent basis (x4 volume => GEH x2). Both "
                           "figures are shown above rather than picking one."))
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note.row, start_column=1, end_row=note.row, end_column=2)
    ws.row_dimensions[note.row].height = 45
    return ws


def write_overview_tab(wb, stats, n_pairs):
    ws = wb.create_sheet("Overview", 0)
    ws.column_dimensions["A"].width = 100

    title = ws.cell(row=1, column=1,
                     value="Vision Camera TMC Accuracy: Baseline vs. 1st Calibration")
    title.font = Font(bold=True, size=14)

    paragraphs = [
        ("Project goal", (
            "Hennepin County's Transportation Operations Department is rolling out a Vision Camera system "
            "that automatically generates turning-movement counts (TMC) at signalized intersections — a task "
            "traditionally done by a person standing at the intersection with a tally sheet. Before the county "
            "can rely on the Vision Camera's counts, it needs to know how accurate they actually are. This "
            "project's goal was to answer that directly: manually count vehicles at a sample of intersections "
            "and compare those manual counts against what the Vision Camera recorded for the same time window."
        )),
        ("Methodology", (
            "Manual counts were collected in 15-minute windows at each intersection and compared to the Vision "
            "Camera's count for that same window, using two standard traffic-engineering accuracy metrics:\n\n"
            "% Difference — the plain relative error, (Vision - Manual) / Manual. Simple to read, but treats a "
            "2-vehicle error the same whether the intersection carries 10 vehicles or 1,000.\n\n"
            "GEH — a chi-squared-derived statistic used industry-wide (state DOTs, FHWA) to flag meaningful "
            "count discrepancies. Formula: GEH = sqrt(2*(Vision-Manual)^2 / (Vision+Manual)). It weights the "
            "error by count magnitude, so it isn't fooled by a large absolute gap on a high-volume intersection "
            "or alarmed by a small absolute gap on a low-volume one. Industry convention treats GEH < 5 (on "
            "hourly volumes) as a good match."
        )),
        ("Scope of this analysis", (
            f"Of the 236 intersections tracked in the county's TMC workbook, manual counts were collected and "
            f"compared for {int(stats['Baseline']['n_intersections'])} intersections in the Baseline round and "
            f"{int(stats['1st Calibration']['n_intersections'])} intersections after the 1st Calibration "
            f"adjustment. {n_pairs} of those intersections were measured in both rounds, which lets us look "
            f"directly at whether calibration improved accuracy at the same locations (see the 'Paired "
            f"Before-After' tab), in addition to the overall comparison across all measured intersections "
            f"(see 'Comparison')."
        )),
        ("A note on 15-minute counts vs. the GEH<5 threshold", (
            "These are 15-minute samples, not hourly volumes. GEH scales with sqrt(volume) for a fixed relative "
            "error, so the conventional GEH<5 threshold — calibrated against hourly counts — is not directly "
            "applicable to raw 15-minute totals: scaling the same counts to an hourly-equivalent basis (x4 "
            "volume) roughly doubles every GEH value. Stats tabs report both the as-collected GEH and the "
            "hourly-equivalent GEH pass rate, rather than applying the threshold to only one basis."
        )),
    ]

    row = 3
    for heading, body in paragraphs:
        h = ws.cell(row=row, column=1, value=heading)
        h.font = Font(bold=True, size=11)
        row += 1
        b = ws.cell(row=row, column=1, value=body)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        n_lines = body.count("\n") + 1 + len(body) // 100
        ws.row_dimensions[row].height = max(30, n_lines * 14)
        row += 2

    ws.sheet_view.showGridLines = False
    return ws


def write_paired_tab(wb, baseline_df, calib_df, top_n=5):
    merged = baseline_df.merge(
        calib_df, on=["ID", "Intersection"], suffixes=("_Baseline", "_Calib")
    ).sort_values("GEH_15min_Baseline", ascending=False).reset_index(drop=True)

    ws = wb.create_sheet("Paired Before-After", 2)
    ws.column_dimensions["A"].width = 45
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15

    title = ws.cell(row=1, column=1,
                     value="Progress at repeated intersections (measured in both Baseline and 1st Calibration)")
    title.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    note = ws.cell(row=2, column=1, value=(
        f"{len(merged)} intersections have comparisons recorded in both rounds. Sorted below by Baseline GEH "
        f"(highest first), so the top {top_n} rows are the intersections that most needed calibration — the "
        f"chart shows the change for those {top_n}; the full table below has all {len(merged)}."
    ))
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 28

    header_row = 4
    headers = ["Intersection", "Baseline GEH", "Calib. GEH", "Baseline % Diff", "Calib. % Diff"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header(ws, len(headers), row=header_row)

    for i, r in merged.iterrows():
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=r["Intersection"])
        ws.cell(row=row, column=2, value=r["GEH_15min_Baseline"]).number_format = "0.00"
        ws.cell(row=row, column=3, value=r["GEH_15min_Calib"]).number_format = "0.00"
        ws.cell(row=row, column=4, value=r["Pct_Diff_Baseline"]).number_format = "0.0%"
        ws.cell(row=row, column=5, value=r["Pct_Diff_Calib"]).number_format = "0.0%"

    last_row = header_row + len(merged)
    top_last_row = header_row + top_n

    geh_chart = BarChart()
    geh_chart.type = "col"
    geh_chart.title = f"GEH: Baseline vs. 1st Calibration (top {top_n} by Baseline GEH)"
    geh_chart.y_axis.title = "GEH"
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=top_last_row)
    cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=top_last_row)
    geh_chart.add_data(data_ref, titles_from_data=True)
    geh_chart.set_categories(cats_ref)
    geh_chart.height, geh_chart.width = 9, 16
    ws.add_chart(geh_chart, f"H{header_row}")

    pct_chart = BarChart()
    pct_chart.type = "col"
    pct_chart.title = f"% Difference: Baseline vs. 1st Calibration (same top {top_n})"
    pct_chart.y_axis.title = "% Difference"
    data_ref2 = Reference(ws, min_col=4, max_col=5, min_row=header_row, max_row=top_last_row)
    pct_chart.add_data(data_ref2, titles_from_data=True)
    pct_chart.set_categories(cats_ref)
    pct_chart.height, pct_chart.width = 9, 16
    ws.add_chart(pct_chart, f"H{header_row + 20}")

    autosize(ws, len(headers))
    return ws


def write_comparison_tab(wb, stats):
    ws = wb.create_sheet("Comparison", 1)

    title_cell = ws.cell(row=1, column=1, value="TMC Vision Camera Accuracy: Baseline vs. 1st Calibration")
    title_cell.font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    header_row = 2
    ws.cell(row=header_row, column=1, value="Metric")
    ws.cell(row=header_row, column=2, value="Baseline")
    ws.cell(row=header_row, column=3, value="1st Calibration")
    style_header(ws, 3, row=header_row)

    # Ordered so the two Mean-GEH rows are adjacent (used directly as chart data below)
    metrics = [
        ("Intersections with data", "n_intersections"),
        ("Mean % Difference", "mean_pct_diff"),
        ("Mean |% Difference|", "mean_abs_pct_diff"),
        ("Mean GEH (15-min, as collected)", "mean_geh_15min"),
        ("Mean GEH (hourly-equivalent)", "mean_geh_hourly"),
        ("Median GEH (15-min, as collected)", "median_geh_15min"),
        ("% Intersections GEH<5 (15-min)", "pct_pass_geh5_15min"),
        ("% Intersections GEH<5 (hourly-equivalent)", "pct_pass_geh5_hourly"),
    ]
    start_row = header_row + 1
    for i, (label, key) in enumerate(metrics):
        r = start_row + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=stats["Baseline"][key])
        ws.cell(row=r, column=3, value=stats["1st Calibration"][key])
        if key in ("mean_pct_diff", "mean_abs_pct_diff"):
            ws.cell(row=r, column=2).number_format = "0.0%"
            ws.cell(row=r, column=3).number_format = "0.0%"
        elif key.startswith("pct_pass"):
            ws.cell(row=r, column=2).number_format = "0.0"
            ws.cell(row=r, column=3).number_format = "0.0"
        elif key.startswith(("mean_geh", "median_geh")):
            ws.cell(row=r, column=2).number_format = "0.00"
            ws.cell(row=r, column=3).number_format = "0.00"

    autosize(ws, 3)

    mean_geh_15min_row = start_row + 3  # index 3 in metrics list
    mean_geh_hourly_row = start_row + 4  # index 4, adjacent

    chart = BarChart()
    chart.type = "col"
    chart.title = "Mean GEH: Baseline vs. 1st Calibration"
    chart.y_axis.title = "GEH"
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=mean_geh_hourly_row)
    cats_ref = Reference(ws, min_col=1, min_row=mean_geh_15min_row, max_row=mean_geh_hourly_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height, chart.width = 9, 16
    ws.add_chart(chart, "F3")

    note = ws.cell(row=start_row + len(metrics) + 2, column=1,
                    value=("Note: counts are 15-minute samples, not hourly volumes. GEH scales with "
                           "sqrt(volume), so scaling to an hourly-equivalent basis (x4 volume) roughly "
                           "doubles GEH. Both the as-collected and hourly-equivalent figures are shown "
                           "since the conventional GEH<5 threshold was calibrated for hourly counts."))
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note.row, start_column=1, end_row=note.row, end_column=3)
    ws.row_dimensions[note.row].height = 45
    return ws


def main():
    with open("tmc_stats.json") as f:
        stats = json.load(f)
    for sheet_name in SHEETS:
        for k in stats[sheet_name]:
            stats[sheet_name][k] = float(stats[sheet_name][k])

    dfs = {sheet_name: pd.read_csv(FILE_MAP[sheet_name]) for sheet_name in SHEETS}
    n_pairs = len(dfs["Baseline"].merge(dfs["1st Calibration"], on=["ID", "Intersection"]))

    wb = Workbook()
    wb.remove(wb.active)

    write_overview_tab(wb, stats, n_pairs)
    write_comparison_tab(wb, stats)
    write_paired_tab(wb, dfs["Baseline"], dfs["1st Calibration"])
    for sheet_name in SHEETS:
        write_data_tab(wb, sheet_name, dfs[sheet_name], stats[sheet_name])
        write_stats_tab(wb, sheet_name, stats[sheet_name])

    out_path = "TMC_Vision_Accuracy_Summary.xlsx"
    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    main()
