"""
Extract Baseline and 1st Calibration % Difference data from TMC_Percent_Difference.xlsx
and compare accuracy progression across the two stages for intersections with data in
both. Kept as a separate script from analyze_tmc.py / consultant_data.py because this
source file has a different shape and is missing information those have:

- One row per intersection per tab (not a 4-row Manual/Vision/%Diff/GEH block).
- The "Total" column already IS the % Difference (Vision vs. Manual), pre-computed --
  there are no raw Manual/Vision counts in this file, so GEH cannot be computed and is
  intentionally not attempted here.
"""

import pandas as pd

SOURCE_FILE = "TMC_Percent_Difference.xlsx"
SHEETS = ["Baseline", "1st Calibration"]
TOTAL_COL = 6  # column F = "Total" (already a % Difference value in this file)


def extract_sheet(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        intersection_id = ws.cell(row=r, column=1).value
        if intersection_id is None:
            continue
        pct_diff = ws.cell(row=r, column=TOTAL_COL).value
        if pct_diff is None or isinstance(pct_diff, str):
            continue  # no data collected for this intersection

        rows.append({
            "ID": intersection_id,
            "Intersection": ws.cell(row=r, column=2).value,
            "Date": ws.cell(row=r, column=4).value,
            "Time of Day": ws.cell(row=r, column=5).value,
            "Pct_Diff": pct_diff,
            "Abs_Pct_Diff": abs(pct_diff),
        })
    return pd.DataFrame(rows)


def build_comparison(baseline_df, calib1_df):
    merged = baseline_df.merge(
        calib1_df, on=["ID", "Intersection"], suffixes=("_Baseline", "_1st_Calibration")
    )
    merged["Improved"] = merged["Abs_Pct_Diff_1st_Calibration"] < merged["Abs_Pct_Diff_Baseline"]
    merged["Change_In_Abs_Pct_Diff"] = (
        merged["Abs_Pct_Diff_1st_Calibration"] - merged["Abs_Pct_Diff_Baseline"]
    )
    return merged.sort_values("Change_In_Abs_Pct_Diff").reset_index(drop=True)


def main():
    import openpyxl
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)

    dfs = {}
    for sheet_name in SHEETS:
        df = extract_sheet(wb[sheet_name])
        dfs[sheet_name] = df
        safe_name = sheet_name.replace(" ", "_")
        df.to_csv(f"pct_diff_{safe_name}.csv", index=False)
        print(f"{sheet_name}: {len(df)} intersections with data -> pct_diff_{safe_name}.csv")

    comparison = build_comparison(dfs["Baseline"], dfs["1st Calibration"])
    comparison.to_csv("pct_diff_comparison.csv", index=False)

    n_improved = comparison["Improved"].sum()
    print(f"\nIntersections present in both tabs: {len(comparison)}")
    print(f"Improved (smaller |% Diff|) from Baseline to 1st Calibration: {n_improved} "
          f"({n_improved / len(comparison) * 100:.1f}%)")
    print(f"Mean |% Diff| Baseline: {comparison['Abs_Pct_Diff_Baseline'].mean():.3f}")
    print(f"Mean |% Diff| 1st Calibration: {comparison['Abs_Pct_Diff_1st_Calibration'].mean():.3f}")
    print("\nSaved pct_diff_comparison.csv")
    print("\nBiggest improvements (Baseline % Diff -> 1st Calibration % Diff):")
    for _, row in comparison.head(5).iterrows():
        print(f"  {row['Intersection']}: {row['Pct_Diff_Baseline']:.3f} -> "
              f"{row['Pct_Diff_1st_Calibration']:.3f}")


if __name__ == "__main__":
    main()
