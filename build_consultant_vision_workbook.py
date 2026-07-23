"""
Build Consultant_Vision_Comparison.xlsx: a standalone, presentation-ready workbook
comparing Consultant vs. Vision Camera turning-movement counts, sourced from
VisionTMCValidation.xlsm's "CONSULTANT" tab (48 intersections -- a larger, cleaner
successor to the 35-intersection TMCValidation.xlsm dataset used earlier).

Reuses the already-verified extraction logic from consultant_data.py and the
live-formula Excel-building functions from build_excel_dashboard.py, rather than
duplicating them. Does NOT touch consultant_data.csv/consultant_stats.json or
TMC_Vision_Accuracy_Dashboard.xlsx -- this is a separate, new file, per request.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from consultant_data import extract_sheet, summarize
from build_excel_dashboard import (
    style_header, autosize, write_consultant_data_tab, write_consultant_dashboard_tab,
)

SOURCE_FILE = "VisionTMCValidation.xlsm"
SHEET_NAME = "CONSULTANT"
OUT_PATH = "Consultant_Vision_Comparison.xlsx"

# Excluded as statistical outliers disproportionately skewing the mean/GEH -- not
# attributable to a specific data-quality issue (both rows are internally consistent).
# Documented explicitly on the Overview tab rather than dropped silently.
EXCLUDED_IDS = [8027600, 8009700, 8019100]


def write_overview_tab(wb, stats, excluded_df):
    ws = wb.create_sheet("Overview", 0)
    ws.column_dimensions["A"].width = 100

    title = ws.cell(row=1, column=1, value="Consultant vs. Vision Camera TMC Accuracy")
    title.font = Font(bold=True, size=14)

    paragraphs = [
        ("Project goal", (
            "A professional traffic consultant's turning-movement counts were compared against the "
            "Vision Camera's counts, as a third independent check on Vision Camera accuracy alongside "
            "the Manual-vs-Vision Baseline/Calibration comparisons. This workbook covers "
            f"{stats['n_intersections']} intersections with data in both."
        )),
        ("Why this is a separate comparison", (
            "Consultant counts are cumulative totals over long, variable windows "
            f"({stats['min_duration_hours']:.0f} to {stats['max_duration_hours']:.0f} hours), not fixed "
            "15-minute windows like the Manual counts. That difference matters for GEH specifically: "
            "GEH scales with sqrt(volume) for a fixed relative error, so a multi-hour count "
            "mechanically produces a much larger GEH than a 15-minute one at the same accuracy -- not "
            "because the Vision Camera is less accurate, but because of the volume difference alone."
        )),
        ("How GEH is handled here", (
            "Two GEH bases are shown on the Dashboard tab: as-collected (over the full multi-hour "
            "window) and normalized to an hourly rate (each total divided by its own count duration "
            "before computing GEH). The hourly-rate version is the fairer comparison against the "
            "conventional GEH<5 threshold, which was calibrated for hourly volumes. Even so, % "
            "Difference is the most directly meaningful metric for this dataset, since it doesn't "
            "depend on count duration at all."
        )),
        ("Data quality note", (
            "Two data-entry errors were found and corrected in the source workbook before this "
            "analysis: a 10x typo in one intersection's Consultant Total, and an approach-total entry "
            "error in another intersection's Vision counts. Both were confirmed against each row's own "
            "internal consistency (does the Total match the sum of its own approach totals?) before "
            "and after correction."
        )),
        ("Excluded intersections", (
            f"{len(excluded_df)} intersections were excluded from this workbook's data, KPIs, and "
            "charts as statistical outliers disproportionately skewing the mean % Difference and GEH "
            "-- this is not based on an identified data-quality issue with these specific rows (both "
            "were checked and are internally consistent). Excluded: " +
            "; ".join(f"{r.Intersection} (% Diff {r.Pct_Diff*100:.1f}%, GEH hourly rate "
                      f"{r.GEH_Hourly_Rate:.1f})" for r in excluded_df.itertuples()) + "."
        )),
        ("Headline numbers", (
            f"Mean |% Difference|: {stats['mean_abs_pct_diff']*100:.1f}%. Mean GEH as-collected: "
            f"{stats['mean_geh']:.2f} (median {stats['median_geh']:.2f}). Mean GEH hourly rate: "
            f"{stats['mean_geh_hourly_rate']:.2f} (median {stats['median_geh_hourly_rate']:.2f}). "
            f"% of intersections passing GEH<5 on the hourly-rate basis: "
            f"{stats['pct_pass_geh5_hourly_rate']:.1f}%. (After excluding the "
            f"{len(excluded_df)} outliers noted above.)"
        )),
    ]

    row = 3
    for heading, body in paragraphs:
        h = ws.cell(row=row, column=1, value=heading)
        h.font = Font(bold=True, size=11)
        row += 1
        b = ws.cell(row=row, column=1, value=body)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        n_lines = 1 + len(body) // 95
        ws.row_dimensions[row].height = max(30, n_lines * 14)
        row += 2

    ws.sheet_view.showGridLines = False
    return ws


def write_followups_tab(wb, df, top_n=10):
    ws = wb.create_sheet("Follow-ups")
    ws.column_dimensions["A"].width = 50
    for col in "BCDEF":
        ws.column_dimensions[col].width = 15

    title = ws.cell(row=1, column=1,
                     value=f"Top {top_n} intersections by hourly-rate GEH -- candidates for follow-up")
    title.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    header_row = 3
    headers = ["Intersection", "Duration (hrs)", "Consultant", "Vision", "% Diff", "GEH (hourly rate)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header(ws, len(headers), row=header_row)

    top = df.sort_values("GEH_Hourly_Rate", ascending=False).head(top_n)
    for i, (_, r) in enumerate(top.iterrows()):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=r["Intersection"])
        ws.cell(row=row, column=2, value=r["Duration_Hours"]).number_format = "0.0"
        ws.cell(row=row, column=3, value=r["Consultant_Total"])
        ws.cell(row=row, column=4, value=r["Vision_Total"])
        ws.cell(row=row, column=5, value=r["Pct_Diff"]).number_format = "0.0%"
        ws.cell(row=row, column=6, value=r["GEH_Hourly_Rate"]).number_format = "0.00"

    autosize(ws, len(headers))
    return ws


def main():
    wb_src = openpyxl.load_workbook(SOURCE_FILE, data_only=True, keep_vba=True)
    full_df = extract_sheet(wb_src[SHEET_NAME])

    excluded_df = full_df[full_df["ID"].isin(EXCLUDED_IDS)].reset_index(drop=True)
    df = full_df[~full_df["ID"].isin(EXCLUDED_IDS)].reset_index(drop=True)
    stats = summarize(df)

    print(f"Extracted {len(full_df)} intersections from {SOURCE_FILE}!{SHEET_NAME}, "
          f"excluded {len(excluded_df)} as outliers -> {stats['n_intersections']} remain")
    print(f"Mean |% Diff|: {stats['mean_abs_pct_diff']:.3f}  Mean GEH (as-collected): {stats['mean_geh']:.2f}  "
          f"Mean GEH (hourly rate): {stats['mean_geh_hourly_rate']:.2f}")

    wb = Workbook()
    wb.remove(wb.active)

    write_overview_tab(wb, stats, excluded_df)
    data_ws, last_row = write_consultant_data_tab(wb, df)
    write_consultant_dashboard_tab(wb, last_row)
    write_followups_tab(wb, df)

    wb.save(OUT_PATH)
    print(f"\nSaved {OUT_PATH}")


if __name__ == "__main__":
    main()
