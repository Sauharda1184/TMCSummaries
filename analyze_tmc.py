"""
Extract Manual vs. Vision Camera turning-movement-count (TMC) totals from
NewBaseline.xlsm ("Baseline" and "1st Calibration" tabs), compute accuracy
metrics (% Difference, GEH) across all intersections that have data, and
write summary outputs (Excel workbook + data for the HTML report).
"""

import math
import json
import openpyxl
import pandas as pd

SOURCE_FILE = "NewBaseline.xlsm"
SHEETS = ["Baseline", "1st Calibration"]
BLOCK_SIZE = 4  # Manual, Vision, % Difference, GEH rows per intersection
TOTAL_COL = 6   # column F = "Total"


def geh(manual, vision):
    if manual + vision == 0:
        return 0.0
    return math.sqrt(2 * (vision - manual) ** 2 / (manual + vision))


def extract_sheet(ws):
    rows = []
    for r in range(2, ws.max_row + 1, BLOCK_SIZE):
        intersection_id = ws.cell(row=r, column=1).value
        if intersection_id is None:
            continue
        intersection = ws.cell(row=r, column=2).value
        date = ws.cell(row=r, column=4).value
        time_of_day = ws.cell(row=r, column=5).value
        manual_total = ws.cell(row=r, column=TOTAL_COL).value
        vision_total = ws.cell(row=r + 1, column=TOTAL_COL).value

        if manual_total is None or vision_total is None:
            continue  # no data collected for this intersection

        pct_diff = (vision_total - manual_total) / manual_total if manual_total else None
        geh_15min = geh(manual_total, vision_total)
        geh_hourly_equiv = geh(manual_total * 4, vision_total * 4)

        rows.append({
            "ID": intersection_id,
            "Intersection": intersection,
            "Date": date,
            "Time of Day": str(time_of_day) if time_of_day else None,
            "Manual_Total": manual_total,
            "Vision_Total": vision_total,
            "Pct_Diff": pct_diff,
            "GEH_15min": geh_15min,
            "GEH_Hourly_Equiv": geh_hourly_equiv,
        })
    return pd.DataFrame(rows)


def summarize(df):
    return {
        "n_intersections": len(df),
        "mean_pct_diff": df["Pct_Diff"].mean(),
        "median_pct_diff": df["Pct_Diff"].median(),
        "mean_abs_pct_diff": df["Pct_Diff"].abs().mean(),
        "mean_geh_15min": df["GEH_15min"].mean(),
        "median_geh_15min": df["GEH_15min"].median(),
        "mean_geh_hourly": df["GEH_Hourly_Equiv"].mean(),
        "median_geh_hourly": df["GEH_Hourly_Equiv"].median(),
        "pct_pass_geh5_15min": (df["GEH_15min"] < 5).mean() * 100,
        "pct_pass_geh5_hourly": (df["GEH_Hourly_Equiv"] < 5).mean() * 100,
    }


def main():
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True, keep_vba=True)

    data = {}
    stats = {}
    for sheet_name in SHEETS:
        df = extract_sheet(wb[sheet_name])
        df = df.sort_values("GEH_15min", ascending=False).reset_index(drop=True)
        data[sheet_name] = df
        stats[sheet_name] = summarize(df)

    # --- Spot-check recomputed values against the workbook's own stored values ---
    ws = wb["Baseline"]
    print("Spot-check (Baseline) — recomputed vs. workbook-stored:")
    for r, label in [(2, "ID 8059100"), (22, "ID 8052100")]:
        manual = ws.cell(row=r, column=TOTAL_COL).value
        vision = ws.cell(row=r + 1, column=TOTAL_COL).value
        stored_pct = ws.cell(row=r + 2, column=TOTAL_COL).value
        stored_geh = ws.cell(row=r + 3, column=TOTAL_COL).value
        print(f"  {label}: pct_diff recomputed={((vision-manual)/manual):.4f} stored={stored_pct:.4f} | "
              f"geh recomputed={geh(manual, vision):.4f} stored={stored_geh:.4f}")

    print()
    for sheet_name in SHEETS:
        print(f"=== {sheet_name} ===")
        print(f"  n intersections with data: {stats[sheet_name]['n_intersections']}")
        for k, v in stats[sheet_name].items():
            if k != "n_intersections":
                print(f"  {k}: {v:.3f}")
        print()

    # Save tidy extracts for reuse (Excel export step + HTML report step)
    for sheet_name in SHEETS:
        safe_name = sheet_name.replace(" ", "_").replace("1st", "1st")
        data[sheet_name].to_csv(f"tmc_{safe_name}.csv", index=False)

    with open("tmc_stats.json", "w") as f:
        json.dump(stats, f, indent=2, default=str)

    print("Saved tidy CSVs and tmc_stats.json")


if __name__ == "__main__":
    main()
