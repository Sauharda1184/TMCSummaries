"""
Build TMC_Vision_Accuracy_Dashboard.xlsx: a version of the summary workbook where
the data lives in a real Excel Table and every derived metric (% Difference, GEH,
GEH hourly-equivalent) and every KPI (means, pass rates) is a live formula over
that table, rather than a value pre-computed in Python. Edit a Manual/Vision count
in the table and the % Diff/GEH for that row, the KPI tiles, and any chart bound to
that row's data all recalculate automatically in Excel.

Native PivotTables/Slicers are not built here: doing so requires hand-authoring
several interdependent OOXML parts (pivotCacheDefinition, pivotCacheRecords,
pivotTable, slicer, slicerCache) that openpyxl does not provide a supported API
for, and there is no way to validate the result against real Excel from this
environment. Real Excel Tables support slicers natively, so the workbook is built
so a user can add one themselves: select a cell in TMC_Data, Insert > Slicer >
choose "Round".
"""

import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

FILE_MAP = {"Baseline": "tmc_Baseline.csv", "1st Calibration": "tmc_1st_Calibration.csv"}
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

TABLE_COLS = ["Round", "ID", "Intersection", "Date", "Time of Day",
              "Manual_Total", "Vision_Total", "Pct_Diff", "Abs_Pct_Diff",
              "GEH_15min", "GEH_Hourly_Equiv"]


def geh(manual, vision):
    if manual + vision == 0:
        return 0.0
    return math.sqrt(2 * (vision - manual) ** 2 / (manual + vision))


def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, ncols):
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        max_len = max((len(str(ws.cell(row=r, column=c).value)) for r in range(1, ws.max_row + 1)
                       if ws.cell(row=r, column=c).value is not None), default=10)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


def write_overview_tab(wb, n_baseline, n_calib):
    ws = wb.create_sheet("Overview", 0)
    ws.column_dimensions["A"].width = 100

    title = ws.cell(row=1, column=1, value="Vision Camera TMC Accuracy — Live Dashboard")
    title.font = Font(bold=True, size=14)

    paragraphs = [
        ("What's different about this version", (
            "This workbook keeps only the raw Manual and Vision counts as data. Every other column — % "
            "Difference, GEH, the hourly-equivalent GEH — is a live formula computed from those two counts, "
            "and every KPI on the Dashboard tab (means, pass rates) is a live formula over the whole table. "
            "Edit a Manual or Vision count on the 'Data' tab and the % Diff/GEH for that row, the KPI tiles, "
            "and the charts bound to that row all recalculate automatically — nothing here is a pasted-in "
            "number or a static picture of a chart."
        )),
        ("Project goal", (
            "Hennepin County's Transportation Operations Department uses a Vision Camera system that "
            "automatically generates turning-movement counts (TMC) at signalized intersections — a task "
            "traditionally done by a person standing at the intersection with a tally sheet. This project "
            "validates those counts by manually counting vehicles at a sample of intersections and comparing "
            "the manual counts against what the Vision Camera recorded for the same 15-minute window."
        )),
        ("Metrics", (
            "% Difference = (Vision - Manual) / Manual. GEH = sqrt(2*(Vision-Manual)^2 / (Vision+Manual)) — "
            "a chi-squared-derived statistic used industry-wide to flag meaningful count discrepancies, "
            "weighted by count magnitude. Industry convention treats GEH < 5 (on hourly volumes) as a good "
            "match. Because these are 15-minute counts, GEH_Hourly_Equiv (= GEH_15min * 2, matching the "
            "sqrt(4) scaling from a 4x volume increase) is also computed, since the conventional threshold "
            "was calibrated for hourly data."
        )),
        ("Scope", (
            f"{n_baseline} intersections have Baseline data and {n_calib} have 1st Calibration data (of 236 "
            f"tracked). All rows live in a single Excel Table named TMC_Data on the 'Data' tab, with a "
            f"'Round' column distinguishing Baseline from 1st Calibration — this is what lets a single "
            f"slicer or table filter switch every chart between the two."
        )),
        ("To add a live filter button (slicer)", (
            "Click any cell inside the TMC_Data table on the 'Data' tab, then Insert > Slicer > check "
            "'Round'. That gives you a clickable Baseline / 1st Calibration button that filters the table "
            "and every chart bound to it — this can't be pre-built from outside Excel without risking a "
            "corrupted file, but it's a 2-click action once the table is in front of you."
        )),
    ]

    row = 3
    for heading, body in paragraphs:
        h = ws.cell(row=row, column=1, value=heading)
        h.font = Font(bold=True, size=11)
        row += 1
        b = ws.cell(row=row, column=1, value=body)
        b.alignment = Alignment(wrap_text=True, vertical="top")
        n_lines = 1 + len(body) // 95
        ws.row_dimensions[row].height = max(30, n_lines * 14)
        row += 2

    ws.sheet_view.showGridLines = False
    return ws


def write_data_tab(wb, baseline_df, calib_df):
    ws = wb.create_sheet("Data")
    ws.append(TABLE_COLS)
    style_header(ws, len(TABLE_COLS))

    def rows_for(df, round_name):
        recs = df.to_dict(orient="records")
        recs.sort(key=lambda r: geh(r["Manual_Total"], r["Vision_Total"]), reverse=True)
        return recs

    all_rows = [("Baseline", r) for r in rows_for(baseline_df, "Baseline")] + \
               [("1st Calibration", r) for r in rows_for(calib_df, "1st Calibration")]

    for round_name, r in all_rows:
        row_idx = ws.max_row + 1
        ws.cell(row=row_idx, column=1, value=round_name)
        ws.cell(row=row_idx, column=2, value=r["ID"])
        ws.cell(row=row_idx, column=3, value=r["Intersection"])
        ws.cell(row=row_idx, column=4, value=str(r["Date"]) if r.get("Date") else None)
        ws.cell(row=row_idx, column=5, value=str(r["Time of Day"]) if r.get("Time of Day") else None)
        ws.cell(row=row_idx, column=6, value=r["Manual_Total"])
        ws.cell(row=row_idx, column=7, value=r["Vision_Total"])
        # Plain relative references rather than [@Column] structured refs: the latter
        # is standard Excel syntax, but LibreOffice's recalculation (used below to
        # verify this workbook, since real Excel isn't available here) doesn't
        # reliably evaluate "this row" structured references, so this avoids
        # shipping formulas that were never actually confirmed to compute correctly.
        ws.cell(row=row_idx, column=8, value=f"=(G{row_idx}-F{row_idx})/F{row_idx}")
        ws.cell(row=row_idx, column=9, value=f"=ABS(H{row_idx})")
        ws.cell(row=row_idx, column=10,
                value=f"=SQRT(2*(G{row_idx}-F{row_idx})^2/(G{row_idx}+F{row_idx}))")
        ws.cell(row=row_idx, column=11, value=f"=J{row_idx}*2")

    last_row = ws.max_row
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=8).number_format = "0.0%"
        ws.cell(row=r, column=9).number_format = "0.0%"
        ws.cell(row=r, column=10).number_format = "0.00"
        ws.cell(row=r, column=11).number_format = "0.00"

    table_ref = f"A1:{get_column_letter(len(TABLE_COLS))}{last_row}"
    table = Table(displayName="TMC_Data", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False,
        showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = "A2"
    autosize(ws, len(TABLE_COLS))

    n_baseline = sum(1 for round_name, _ in all_rows if round_name == "Baseline")
    n_calib = sum(1 for round_name, _ in all_rows if round_name == "1st Calibration")
    baseline_block = (2, 1 + n_baseline)
    calib_block = (2 + n_baseline, last_row)
    return ws, baseline_block, calib_block, n_baseline, n_calib


def add_geh_bar_chart(data_ws, dest_ws, title, first_row, last_row, anchor):
    # Deliberately not using titles_from_data: for the 1st-Calibration block, the row
    # above first_row is the last Baseline data row, not a header, so a data-derived
    # title would silently mislabel the series using that row's GEH value.
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = "GEH"
    chart.height, chart.width = 9, 20
    data_ref = Reference(data_ws, min_col=10, min_row=first_row, max_row=last_row)
    cats_ref = Reference(data_ws, min_col=3, min_row=first_row, max_row=last_row)
    series = Series(data_ref, title="GEH")
    chart.series.append(series)
    chart.set_categories(cats_ref)
    chart.legend = None
    dest_ws.add_chart(chart, anchor)


def add_scatter_chart(data_ws, dest_ws, title, first_row, last_row, anchor):
    scatter = ScatterChart()
    scatter.title = title
    scatter.x_axis.title = "Manual Total"
    scatter.y_axis.title = "Vision Total"
    scatter.height, scatter.width = 9, 12
    xvalues = Reference(data_ws, min_col=6, min_row=first_row, max_row=last_row)
    yvalues = Reference(data_ws, min_col=7, min_row=first_row, max_row=last_row)
    series = Series(yvalues, xvalues, title="Intersections")
    series.marker.symbol = "circle"
    series.graphicalProperties.line.noFill = True
    scatter.series.append(series)
    dest_ws.add_chart(scatter, anchor)


def add_all_rows_bar_chart(data_ws, dest_ws, last_row, anchor):
    chart = BarChart()
    chart.title = "All intersections — filter/slicer this one by Round"
    chart.y_axis.title = "GEH"
    chart.height, chart.width = 9, 28
    data_ref = Reference(data_ws, min_col=10, min_row=1, max_row=last_row)
    cats_ref = Reference(data_ws, min_col=3, min_row=2, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None
    dest_ws.add_chart(chart, anchor)


def write_dashboard_tab(wb, baseline_block, calib_block):
    ws = wb.create_sheet("Dashboard", 1)
    ws.column_dimensions["A"].width = 32
    for col in "BC":
        ws.column_dimensions[col].width = 16

    title = ws.cell(row=1, column=1, value="TMC Vision Camera Accuracy — Live KPIs")
    title.font = Font(bold=True, size=13)
    ws.merge_cells("A1:C1")

    ws.cell(row=2, column=1,
            value="Every value below is a formula over the TMC_Data table (Data tab) — "
                  "edit a row there and these update automatically.").alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:C2")
    ws.row_dimensions[2].height = 28

    header_row = 4
    ws.cell(row=header_row, column=1, value="Metric")
    ws.cell(row=header_row, column=2, value="Baseline")
    ws.cell(row=header_row, column=3, value="1st Calibration")
    style_header(ws, 3, row=header_row)

    rows = [
        ("Intersections with data", '=COUNTIFS(TMC_Data[Round],"{r}")', "0"),
        ("Mean % Difference", '=AVERAGEIFS(TMC_Data[Pct_Diff],TMC_Data[Round],"{r}")', "0.0%"),
        ("Mean |% Difference|", '=AVERAGEIFS(TMC_Data[Abs_Pct_Diff],TMC_Data[Round],"{r}")', "0.0%"),
        ("Mean GEH (as-collected)", '=AVERAGEIFS(TMC_Data[GEH_15min],TMC_Data[Round],"{r}")', "0.00"),
        ("Mean GEH (hourly-equivalent)", '=AVERAGEIFS(TMC_Data[GEH_Hourly_Equiv],TMC_Data[Round],"{r}")', "0.00"),
        ("% GEH<5 (as-collected)",
         '=COUNTIFS(TMC_Data[Round],"{r}",TMC_Data[GEH_15min],"<5")/COUNTIFS(TMC_Data[Round],"{r}")', "0.0%"),
        ("% GEH<5 (hourly-equivalent)",
         '=COUNTIFS(TMC_Data[Round],"{r}",TMC_Data[GEH_Hourly_Equiv],"<5")/COUNTIFS(TMC_Data[Round],"{r}")', "0.0%"),
    ]
    r = header_row + 1
    for label, formula, numfmt in rows:
        ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=formula.format(r="Baseline"))
        c3 = ws.cell(row=r, column=3, value=formula.format(r="1st Calibration"))
        c2.number_format = numfmt
        c3.number_format = numfmt
        r += 1
    autosize(ws, 3)

    note = ws.cell(row=r + 1, column=1, value=(
        "Note: counts are 15-minute samples, not hourly volumes, so the conventional GEH<5 threshold "
        "(calibrated for hourly counts) is shown on both bases above rather than applied to only one — "
        "see the Overview tab."))
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note.row, start_column=1, end_row=note.row, end_column=3)
    ws.row_dimensions[note.row].height = 42

    data_ws = wb["Data"]
    add_geh_bar_chart(data_ws, ws, "Baseline: GEH by Intersection (sorted)",
                       baseline_block[0], baseline_block[1], f"E{header_row}")
    add_geh_bar_chart(data_ws, ws, "1st Calibration: GEH by Intersection (sorted)",
                       calib_block[0], calib_block[1], f"E{header_row + 20}")
    add_scatter_chart(data_ws, ws, "Baseline: Manual vs. Vision Total",
                       baseline_block[0], baseline_block[1], f"E{header_row + 40}")
    add_scatter_chart(data_ws, ws, "1st Calibration: Manual vs. Vision Total",
                       calib_block[0], calib_block[1], f"E{header_row + 60}")
    return ws


def write_paired_tab(wb, baseline_df, calib_df, top_n=5):
    merged_ids = sorted(set(baseline_df["ID"]) & set(calib_df["ID"]))
    ws = wb.create_sheet("Paired Before-After")
    ws.column_dimensions["A"].width = 45
    for col in "BCDE":
        ws.column_dimensions[col].width = 15

    title = ws.cell(row=1, column=1,
                     value="Progress at repeated intersections — live lookups from TMC_Data")
    title.font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    note = ws.cell(row=2, column=1, value=(
        f"{len(merged_ids)} intersections appear in both rounds. Each cell below is a SUMPRODUCT lookup "
        f"into TMC_Data by ID and Round, so these also update if the underlying counts change."))
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    ws.row_dimensions[2].height = 28

    header_row = 4
    headers = ["Intersection", "Baseline GEH", "Calib. GEH", "Baseline % Diff", "Calib. % Diff"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header(ws, len(headers), row=header_row)

    id_to_name = dict(zip(baseline_df["ID"], baseline_df["Intersection"]))
    id_to_name.update(dict(zip(calib_df["ID"], calib_df["Intersection"])))
    baseline_geh = {r["ID"]: geh(r["Manual_Total"], r["Vision_Total"]) for r in baseline_df.to_dict("records")}
    sorted_ids = sorted(merged_ids, key=lambda i: baseline_geh.get(i, 0), reverse=True)

    for i, id_val in enumerate(sorted_ids):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=id_to_name[id_val])
        for col, round_name, field in [(2, "Baseline", "GEH_15min"), (3, "1st Calibration", "GEH_15min"),
                                        (4, "Baseline", "Pct_Diff"), (5, "1st Calibration", "Pct_Diff")]:
            formula = (f'=SUMPRODUCT((TMC_Data[ID]={id_val})*(TMC_Data[Round]="{round_name}")'
                       f'*TMC_Data[{field}])')
            cell = ws.cell(row=row, column=col, value=formula)
            cell.number_format = "0.00" if field == "GEH_15min" else "0.0%"

    last_row = header_row + len(sorted_ids)
    top_last_row = header_row + top_n

    geh_chart = BarChart()
    geh_chart.type = "col"
    geh_chart.title = f"GEH: Baseline vs. 1st Calibration (top {top_n} by Baseline GEH)"
    geh_chart.y_axis.title = "GEH"
    data_ref = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=top_last_row)
    cats_ref = Reference(ws, min_col=1, min_row=header_row + 1, max_row=top_last_row)
    geh_chart.add_data(data_ref, titles_from_data=True)
    geh_chart.set_categories(cats_ref)
    geh_chart.height, geh_chart.width = 9, 16
    ws.add_chart(geh_chart, f"G{header_row}")

    autosize(ws, len(headers))
    return ws


def write_all_intersections_tab(wb, last_row_data):
    ws = wb.create_sheet("All Intersections (filterable)")
    ws.cell(row=1, column=1,
            value="Bound to the full TMC_Data table — add a slicer on 'Round' (Data tab: click table, "
                  "Insert > Slicer) to filter this live.")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30
    ws["A1"].alignment = Alignment(wrap_text=True)
    add_all_rows_bar_chart(wb["Data"], ws, last_row_data, "A3")
    return ws


def main():
    baseline_df = pd.read_csv(FILE_MAP["Baseline"])
    calib_df = pd.read_csv(FILE_MAP["1st Calibration"])

    wb = Workbook()
    wb.remove(wb.active)

    data_ws, baseline_block, calib_block, n_baseline, n_calib = write_data_tab(wb, baseline_df, calib_df)
    write_overview_tab(wb, n_baseline, n_calib)
    write_dashboard_tab(wb, baseline_block, calib_block)
    write_paired_tab(wb, baseline_df, calib_df)
    write_all_intersections_tab(wb, data_ws.max_row)

    # Sheet creation order left us [Overview, Dashboard, Data, ...]; swap Data and
    # Dashboard so Data (the source table) reads right after Overview.
    wb.move_sheet("Data", offset=-1)

    out_path = "TMC_Vision_Accuracy_Dashboard.xlsx"
    wb.save(out_path)
    print(f"Saved {out_path}: {n_baseline} Baseline rows, {n_calib} Calibration rows")


if __name__ == "__main__":
    main()
