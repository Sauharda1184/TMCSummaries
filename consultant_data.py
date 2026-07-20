"""
Extract Consultant vs. Vision Camera turning-movement-count (TMC) totals from
TMCValidation.xlsm's "CONSULTANT" tab, compute % Difference and GEH across every
intersection that has data, and save tidy outputs for the report/workbook steps.

Kept separate from analyze_tmc.py (which handles the Manual-vs-Vision Baseline /
1st Calibration comparison) because this is a different comparison group with a
different counting method: Consultant counts are cumulative totals over long,
variable windows (e.g. "6:00 AM to 7:00 PM", 13 hours) rather than fixed 15-minute
windows, so mixing them into the same stats/thresholds as the Manual comparison
would be misleading.
"""

import math
import re
import json
import openpyxl
import pandas as pd

SOURCE_FILE = "TMCValidation.xlsm"
SHEET_NAME = "CONSULTANT"
BLOCK_SIZE = 4  # Manual/Consultant, Vision, % Difference, GEH rows per intersection
TOTAL_COL = 6   # column F = "Total"

TIME_RANGE_RE = re.compile(
    r"(\d{1,2})(?::(\d{2}))?\s*([AP]M)\s*to\s*(\d{1,2})(?::(\d{2}))?\s*([AP]M)",
    re.IGNORECASE,
)


def geh(a, b):
    if a + b == 0:
        return 0.0
    return math.sqrt(2 * (b - a) ** 2 / (a + b))


def parse_duration_hours(time_range):
    if not time_range or not isinstance(time_range, str):
        return None
    m = TIME_RANGE_RE.search(time_range)
    if not m:
        return None
    h1, m1, ap1, h2, m2, ap2 = m.groups()

    def to_24h(h, mins, ampm):
        h = int(h) % 12
        if ampm.upper() == "PM":
            h += 12
        mins = int(mins) if mins else 0
        return h + mins / 60

    start, end = to_24h(h1, m1, ap1), to_24h(h2, m2, ap2)
    duration = end - start
    if duration < 0:
        duration += 24  # window crosses midnight
    return round(duration, 2)


def extract_sheet(ws):
    rows = []
    for r in range(2, ws.max_row + 1, BLOCK_SIZE):
        intersection_id = ws.cell(row=r, column=1).value
        if intersection_id is None:
            continue
        intersection = ws.cell(row=r, column=2).value
        label = ws.cell(row=r, column=3).value
        date = ws.cell(row=r, column=4).value
        time_of_day = ws.cell(row=r, column=5).value
        consultant_total = ws.cell(row=r, column=TOTAL_COL).value
        vision_total = ws.cell(row=r + 1, column=TOTAL_COL).value

        if consultant_total is None or vision_total is None:
            continue  # no data collected for this intersection (or Date/Time-only rows)

        pct_diff = (vision_total - consultant_total) / consultant_total if consultant_total else None
        geh_val = geh(consultant_total, vision_total)
        duration_hours = parse_duration_hours(time_of_day)
        # Normalize to an hourly rate before computing GEH, so the result is on the
        # same basis the conventional GEH<5 threshold was actually calibrated for
        # (hourly volumes) -- unlike the raw GEH above, which mechanically inflates
        # with sqrt(volume) for a window this long regardless of true accuracy.
        geh_hourly_rate = (geh(consultant_total / duration_hours, vision_total / duration_hours)
                            if duration_hours else None)

        rows.append({
            "ID": intersection_id,
            "Intersection": intersection,
            "Label": str(label).strip() if label else None,  # normalizes "Consultant "/"Consultant"
            "Date": date,
            "Time of Day": time_of_day,
            "Duration_Hours": duration_hours,
            "Consultant_Total": consultant_total,
            "Vision_Total": vision_total,
            "Pct_Diff": pct_diff,
            "GEH": geh_val,
            "GEH_Hourly_Rate": geh_hourly_rate,
        })
    return pd.DataFrame(rows)


def summarize(df):
    return {
        "n_intersections": len(df),
        "mean_pct_diff": df["Pct_Diff"].mean(),
        "median_pct_diff": df["Pct_Diff"].median(),
        "mean_abs_pct_diff": df["Pct_Diff"].abs().mean(),
        "mean_geh": df["GEH"].mean(),
        "median_geh": df["GEH"].median(),
        "pct_pass_geh5": (df["GEH"] < 5).mean() * 100,
        "mean_geh_hourly_rate": df["GEH_Hourly_Rate"].mean(),
        "median_geh_hourly_rate": df["GEH_Hourly_Rate"].median(),
        "pct_pass_geh5_hourly_rate": (df["GEH_Hourly_Rate"] < 5).mean() * 100,
        "mean_duration_hours": df["Duration_Hours"].mean(),
        "min_duration_hours": df["Duration_Hours"].min(),
        "max_duration_hours": df["Duration_Hours"].max(),
    }


def main():
    wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True, keep_vba=True)
    ws = wb[SHEET_NAME]

    df = extract_sheet(ws)
    df = df.sort_values("GEH", ascending=False).reset_index(drop=True)
    stats = summarize(df)

    print("Spot-check — recomputed vs. workbook-stored:")
    spot_check_ids = [8044600, 8057300]
    for target_id in spot_check_ids:
        r = next(r for r in range(2, ws.max_row + 1, BLOCK_SIZE)
                  if ws.cell(row=r, column=1).value == target_id)
        consultant = ws.cell(row=r, column=TOTAL_COL).value
        vision = ws.cell(row=r + 1, column=TOTAL_COL).value
        stored_pct = ws.cell(row=r + 2, column=TOTAL_COL).value
        stored_geh = ws.cell(row=r + 3, column=TOTAL_COL).value
        print(f"  ID {target_id}: pct_diff recomputed={((vision - consultant) / consultant):.4f} "
              f"stored={stored_pct:.4f} | geh recomputed={geh(consultant, vision):.4f} stored={stored_geh:.4f}")

    print(f"\n=== CONSULTANT vs. Vision ===")
    print(f"  n intersections with data: {stats['n_intersections']}")
    for k, v in stats.items():
        if k != "n_intersections":
            print(f"  {k}: {v:.3f}")

    df.to_csv("consultant_data.csv", index=False)
    with open("consultant_stats.json", "w") as f:
        json.dump(stats, f, indent=2, default=str)
    print("\nSaved consultant_data.csv and consultant_stats.json")


if __name__ == "__main__":
    main()
